from openpyxl import load_workbook, Workbook
from pathlib import Path
from user_interface import *

def list_values(row_list): # Создание массива данных новой таблицы
    table_list_values.append(row_list)
    return table_list_values

def filling_data(table_list_values,sheet_new, row = 0):# запись данных в новую таблицу
    for item in table_list_values:
        row += 1
        sheet_new['A' + str(row)] = item[0]
        sheet_new['B' + str(row)] = item[1]
        sheet_new['C' + str(row)] = item[2]


def new_workbook(file_name_new, table_list_values): # создание, заполнение и сохранение рабочей книги
    wb = Workbook()  # новая рабочая книга
    sheet_new = wb.active
    sheet_new.title = file_name_new
    filling_data(table_list_values, sheet_new)  # запись данных в новую таблицу
    wb.save(file_name_new)

file_name = ''
dialog = A()
dialog.dialog_window()
file_name = dialog.file_name
print(file_name)
file_name_new = Path(file_name).stem + '_new' + '.xlsx' # создал имя для нового файла
table_list_values = [] # пустой массив данных для новой таблицы


if len(file_name) > 0:
    wb2 = load_workbook(file_name, data_only=True)  # Загрузил файл в ячейках только значения
    sheet = wb2['Лист1']
    row_number = 0
    for row in sheet.rows: # перебираю файл построчно
        row_number += 1
        list_values([row_number, ' '.join(map(str, [row[1].value, 'Размер', row[3].value, 'штрих код', row[8].value])), row[9].value]) # создал новую строку для записи в ячейку
    new_workbook(file_name_new, table_list_values)
else:
    error_input_file()