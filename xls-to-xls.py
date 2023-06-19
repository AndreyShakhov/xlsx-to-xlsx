from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
def list_values(row_list):# Создание массива данных новой таблицы
    table_list_values.append(row_list)
    return table_list_values

def Filling_data(table_list_values,row = 0):# запись данных в новую таблицу
    for item in table_list_values:
        row += 1
        sheet_new['A' + str(row)] = item[0]
        sheet_new['B' + str(row)] = item[1]
        sheet_new['C' + str(row)] = item[2]

def open_file():
    file_open = filedialog.askopenfilename(filetypes = (("Text files","*.xlsx"),("all files","*.*")))
    print(file_open)

    return file_open

def error_input_file():
    messagebox.showinfo('Ошибка', 'Вы не выбрали файл \nЗапустите программу заново')

def dialog_window():
    window = Tk()
    window.title("Сложение столбцов")
    window.geometry('400x250')
    btn = Button(window, text="Открыть файл", command=open_file)
    btn.grid(column=1, row=1)
    window.mainloop()
    return file_open_1

file_open_1= open_file()

dialog_window()
#open_file()
file_name = 'Спецификация.xlsx'
file_name_new = Path(file_name).stem + '_new' + '.xlsx' # создал имя для нового файла
table_list_values = [] # пустой массив данных для новой таблицы
#print(file_open)

row_number = 0

wb = Workbook() # новая рабочая книга
sheet_new = wb.active
sheet_new.title = file_name_new


print(file_open, 'file_open')
if len(file_open) > 0:
    wb2 = load_workbook('Спецификация.xlsx', data_only=True)  # Загрузил файл в ячейках только значения
    sheet = wb2['Лист1']
    for row in sheet.rows: # перебираю файл построчно
        row_number += 1
        new_line = ' '.join(map(str, [row[1].value, 'Размер', row[3].value, 'штрих код', row[8].value]))# создал новую строку для записи в ячейку
        list_values([row_number, ' '.join(map(str, [row[1].value, 'Размер', row[3].value, 'штрих код', row[8].value])), row[9].value])

    Filling_data(table_list_values)
    wb.save(file_name_new)
else:
    error_input_file()



